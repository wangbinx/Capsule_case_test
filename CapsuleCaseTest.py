import os
import copy
import shutil
import subprocess
import openpyxl
import openpyxl.styles as styles
import argparse

result_dict = {}
LogMsg = {}

class Parse_command(object):

	def __init__(self, File):
		self.file = File
		self.Command_List = self._InQuote()
		self.value_command = ['-o','--output', '--capoemflag', '--guid', '--hardware-instance', '--monotonic-count',
		                      '--fw-version', '--lsv', '--signing-tool-path', '--debug']
		self.mode_command = ['-e','--encode', '-d','--decode', '--dump-info']
		self.others_command = ['-h','--help', '-v','--verbose', '-q','--quiet', '--version']
		self.cert_command = ['--pfx-file', '--signer-private-cert', '--other-public-cert', '--trusted-public-cert']
		self.capflag_command = ['--capflag']

	def _read(self):
		try:
			with open(self.file,'r') as F:
				lines =F.readlines()
			if lines:
				if len(lines) > 1:
					command = ''.join(lines)
				else:
					command = lines[0]
			else:
				print("No command in CommandFile")
				command = ''
			return command
		except Exception as err:
			print("Not Command.txt found")
			print(err)

	#Parse command with quote
	def _InQuote(self):
		List = self._read().split(' ')
		if '"' in self._read():
			for i in List:
				if i.startswith('"'):
					start = List.index(i)
				if i.endswith('"'):
					end = List.index(i)
			try:
				str = ' '.join(List[start:end + 1])
			except Exception as e:
				print(e)
				print("Parse Command Error: only one '\"' in command")
				sys.exit(1)
			List = List[:start]
			List.append(str)
		List = [i for i in List if i != '']
		return List

	#Parse input file
	def _input(self):
		tmp = copy.copy(self.Command_List)
		for i in range(len(tmp)-1, -1, -1):
			if tmp[i] in self.value_command + self.cert_command + self.capflag_command:
				tmp.pop(-1)
				tmp.pop(i)
			elif tmp[i] in self.mode_command + self.others_command:
				tmp.pop(i)
		input = tmp[0]
		return input

	#Parse CapFlag
	def capflag(self):
		Flag = []
		for i in self.Command_List:
			if i in self.capflag_command:
				index = self.Command_List.index(i)
				Flag.append(self.Command_List[index + 1])
		return Flag


	def mode(self):
		for i in self.Command_List:
			if i in self.mode_command:
				return i
		return

	def value_dict(self):
		Value = {}
		for i in self.Command_List:
			if i in self.value_command + self.cert_command:
				if self.Command_List.index(i) + 1 < len(self.Command_List):
					if self.Command_List[self.Command_List.index(i) + 1] not in self.value_command + self.cert_command + self.capflag_command+self.mode_command + self.others_command:
						Value[i] = self.Command_List[self.Command_List.index(i) + 1]
				else:
					Value[i] =""
		capflag = self.capflag()
		if capflag:
			Value['--capflag'] = capflag
		run_mode = self.mode()
		if run_mode:
			Value['mode'] = run_mode
		return Value

	def output(self):
		if "-o" in self.value_dict().keys():
			return self.value_dict()["-o"]
		elif "--output" in self.value_dict().keys():
			return self.value_dict()["--output"]
		return

	def command_str(self):
		return self._read()

class Run(Parse_command):

	def __init__(self,casepath):
		self.CFileName = 'command.txt'
		self.casepath = casepath
		self.CFile = os.path.join(self.casepath, self.CFileName)
		super(Run, self).__init__(self.CFile)
		self.root = os.getcwd()
		self.script = "GenerateCapsule.py"
		self.ExpectedResult = os.path.join(self.casepath, "ExpectedResult")
		self.resultpath = os.path.join(self.casepath, "TestResult")
		self.case = os.path.basename(casepath)
		self.reportcaseid = self.case.replace("TC","Case")
		self.LOG =''

	def process(self):
		print("Running case:%s"%self.case)
		InputFlag = False
		for root, dirs, files in os.walk(self.casepath, topdown=True, followlinks=False):
			if "InputFile" in dirs:
				InputFlag = True
			for name in files:
				if os.path.splitext(name)[-1].lower() == '.bat':
					self._runbat(root,os.path.join(root,name))
		files = self.MoveInputFile(self.casepath,InputFlag)
		#Run Script, Create run log
		run = subprocess.Popen('python %s %s' % (self.script,self.command_str()),stdout=subprocess.PIPE, stderr=subprocess.PIPE,shell=True)
		out = run.stdout.read()
		err = run.stderr.read()
		if out !=b"" and out !="":
			self.write2log("result.log",out)
		elif err != "" and err != b"":
			self.write2log("result.log", err)
		else:
			self.write2log("result.log",out+err)
		#Replace result file and clean root dir
		if os.path.exists(self.resultpath):
			shutil.rmtree(self.resultpath)
		os.makedirs(self.resultpath)
		shutil.move("result.log",self.resultpath)
		if self.output() and os.path.exists(self.output()):
			shutil.move(self.output(),self.resultpath)
		#Compare file to Get result
		testresult = self.Result(self.case)
		result_dict[self.reportcaseid] = testresult
		if InputFlag:
			try:
				self._removeinput(files)
			except Exception as err:
				print(err)

	# Copy InputFile to root dir
	def MoveInputFile(self,Path,Flag):
		files = []
		if Flag:
			for root, dirs, File in os.walk(os.path.join(Path,"InputFile"), topdown=True, followlinks=False):
				for file in File:
					files.append(file)
					shutil.copy(os.path.join(root,file),self.root)
		return files

	#Read Command File to get OutputFile name
	def GetOutputFileName(self,CFile):
		output = ''
		command = ''
		with open(CFile,'r') as CF:
			lines = CF.readlines()
		for c in lines:
			List = c.split(' ')
			if "-o" in List:
				output = List[List.index('-o')+1].strip()
			command =c.strip()
		return output, command

	def write2log(self,filename,string):
		with open(filename,'wb') as f:
			f.write(string)
	#Remove InputFile
	def _removeinput(self,files):
		for i in files:
			os.remove(os.path.join(self.root, i))

	# Run bat file if have bat in folder
	def _runbat(self,path,file):
		os.chdir(path)
		result = subprocess.Popen(file,shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
		os.chdir(self.root)
		out = result.stdout.read()
		err = result.stderr.read()
		if err == '':
			print("%s %s"%(self.case,out))
			self.LOG +="%s %s"%(self.case,out)
			return True
		else:
			print("%s%s"%(self.case,err))
			self.LOG +="%s%s"%(self.case,err)

	def Result(self,caseid):
		FileResult = True
		LogName ='result.log'
		exlog =os.path.join(self.ExpectedResult,LogName)
		testlog = os.path.join(self.resultpath,LogName)
		LogResult = self.comparelog(exlog,testlog)
		if self.output() != None:
			exfile = os.path.join(self.ExpectedResult,self.output())
			testfile = os.path.join(self.resultpath,self.output())
			if os.path.exists(exfile):
				if os.path.exists(testfile) and os.stat(testfile).st_size != 0:
					if self.mode() == "-e" or self.mode() == "--encode":
						FileResult =self._comparebin(exfile,testfile)
					elif self.mode() == "-d" or self.mode() == "--decode":
						FileResult =self._comparedecodeoutput(exfile,testfile)
				else:
					FileResult = False
				print("%s File Compare result:%s,"%(caseid,self._convert(FileResult))),
				self.LOG += "%s File Compare result:%s; "%(caseid,self._convert(FileResult))
		print("%s Log Compare result:%s," % (caseid,self._convert(LogResult))),
		self.LOG += "%s Log Compare result:%s; " % (caseid, self._convert(LogResult))
		result =self._convert(LogResult&FileResult)
		self.write_result_to_File(caseid,result,self.resultpath)
		print("Test result is:%s"%(result))
		self.LOG += "%s Test result is:%s\n\n" % (caseid, result)
		LogMsg[int(caseid.split("-")[1])] = self.LOG
		return result

	def _convert(self,result):
		if result == True:
			return "PASS"
		elif result == False:
			return "FAIL"

	def write_result_to_File(self,id,result,path):
		with open(os.path.join(path,"%s,%s"%(id,result)),'w+') as R:
			R.write("%s test result is %s"%(id,result))

	def _comparedecodeoutput(self,file1,file2):
		with open(file1,'r')as f1:
			read1 = f1.read()
		with open(file2,'r')as f2:
			read2 = f2.read()
		if read1 == read2:
			return True
		return False

	#Compare result.log
	def comparelog(self,ori,test):
		result = False
		with open(ori,'r')as of:
			ori_read = of.read()
		with open(test,'r') as tf:
			test_read = tf.read()
		if self._removespace(ori_read) == self._removespace(test_read):
			result = True
		return result

	def _removespace(self,string):
		string =string.replace(" ",'').replace('\r','').replace("\n",'')
		return string

	def _payloadsize(self,file):
		stat = os.stat(file)
		return stat.st_size

	#Read bin file and return
	def _readbin(self,file,inputfile):
		with open(file, 'rb') as f:
			all = f.read()
		header = all[:0x5f]
		payloadsize = self._payloadsize(inputfile)
		fpm = all[-(0x10+payloadsize):-payloadsize]
		return header, fpm

	#Copmare Bin file and return result
	def _comparebin(self,file1,file2):
		#inputfile = os.path.join(self.casepath,"InputFile",self._input())
		result = False
		f1f, f1l = self._readbin(file1,self._input())
		f2f, f2l = self._readbin(file2,self._input())
		if f1f == f2f:
			if f1l == f2l:
				result = True
		return result



class Excel(object):

	def __init__(self,reportfile,Dict):
		self.report = openpyxl.load_workbook(reportfile)
		self.Result_Dict = Dict
		active = self.report.active
		self.sheet = self.report[active.title]
		self.border = styles.Border(left=styles.Side(style='thin', color='FF000000'), right=styles.Side(style='thin', color='FF000000'),
		       top=styles.Side(style='thin', color='FF000000'), bottom=styles.Side(style='thin', color='FF000000'),
		       diagonal=styles.Side(style='thin', color='FF000000'), diagonal_direction=0,
		       outline=styles.Side(style='thin', color='FF000000'), vertical=styles.Side(style='thin', color='FF000000'),
		       horizontal=styles.Side(style='thin', color='FF000000'))

	def write_to_excel(self):
		ID_col = "E"
		Rst_col = "F"
		for title in self.sheet[1]:
			if title.value == 'ID':
				ID_col = title.column
			elif title.value == 'Result':
				Rst_col = title.column
		for n in range(2,self.sheet.max_row+1):
			if self.sheet['%s%d'%(ID_col,n)].value:
				if self.sheet['%s%d'%(ID_col,n)].value in self.Result_Dict.keys():
					self.sheet['%s%d' % (Rst_col, n)].value = self.Result_Dict[self.sheet['%s%d'%(ID_col,n)].value]
					self.sheet['%s%d' % (Rst_col, n)].font = self.color(self.sheet['%s%d' % (Rst_col, n)].value)

	def addborder(self):
		for n in range(1,self.sheet.max_row+1):
			for m in range(1,self.sheet.max_column+1):
				self.sheet.cell(row=n,column=m).border = self.border
		print("Fix excel border finish")

	def save(self):
		self.write_to_excel()
		self.addborder()
		print("Save result to test.xlsx")
		return self.report.save(os.path.join(os.getcwd(),'test.xlsx'))

	def color(self,value):
		if value == "FAIL":
			return styles.Font(color=styles.colors.RED)
		elif value == "PASS":
			return styles.Font(color=styles.colors.GREEN)
		else:
			return

def FindCasePath(Path):
	PATH = {}
	for root, dirs, File in os.walk(Path, topdown=True, followlinks=False):
		for dir in dirs:
			if "TC-" in dir:
				casepath = os.path.join(root, dir)
				PATH[dir] = casepath
	return PATH

def main(Path):
	CaseDict = FindCasePath(Path)
	parser = argparse.ArgumentParser()
	parser.add_argument('-id', '--caseid', metavar="", dest='caseid', help="Input case ID.")
	parser.add_argument('-f', '--folder', metavar="", dest='folder', help="Input folder name(e.g. Testcase\Basic).")
	#parser.add_argument('-c', '--classify', metavar="", dest='classify', help="Input case classify(e.g. Basic/Encode).")
	options = parser.parse_args()
	if options.caseid:
		r = Run(CaseDict[options.caseid])
		r.process()
	elif options.folder:
		for root, dirs, File in os.walk(options.folder, topdown=True, followlinks=False):
			for dir in dirs:
				if "TC-" in dir:
					casepath = os.path.join(root,dir)
					r = Run(casepath)
					r.process()
		Excel("report.xlsx",result_dict).save()
	else:
		for root, dirs, File in os.walk(Path, topdown=True, followlinks=False):
			for dir in dirs:
				if "TC-" in dir:
					casepath = os.path.join(root,dir)
					r = Run(casepath)
					r.process()
		Excel("report.xlsx",result_dict).save()
	with open("log.txt", 'w+') as log:
		aa = list(LogMsg.keys())
		aa.sort()
		for key in aa:
			log.write(LogMsg[key])

if __name__ == "__main__":
	root = os.getcwd()
	os.environ["PYTHONPATH"] = os.path.dirname(root)
	main(os.path.join(os.getcwd(),'Testcase'))